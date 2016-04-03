#!/usr/bin/env python
# -*- coding: utf-8 -*-
""" HikikakuKunSheets: extract game strategies of Shogi games
   and put Excel Sheets for viewing shogi games in convenient ways."""

import sys
import os
import os.path
import codecs
import re
import datetime
from openpyxl import Workbook
from openpyxl.cell import get_column_letter


def retrieve_files(directly):
    """ retrieve csa and kif files.
    """
    new_files = []
    csa_files = []

    for root, dirs, files in os.walk(directly):
        dirs = dirs
        for fil in files:
            path = os.path.join(root, fil)
            if fil.endswith('.kif'):
                new_files += [path]
                # print "kif:%s" % path
            elif fil.endswith('.csa'):
                csa_files += [path]
                # print "csa:%s" % path

    remove_list = []
    csa_new_files = []
    for cfi in csa_files:
        root, ext = os.path.splitext(cfi)
        ext = ext
        kiffile = root + '.kif'

        if not os.path.isfile(kiffile):
            print 'remove: %s' % kiffile
            remove_list += [cfi]
        else:
            csa_new_files += [cfi]

    # print "newFiles = %d" % len(newFiles)
    # print "csaFiles = %d" % len(csaFiles)
    # print "csaNewFiles = %d" % len(csaNewFiles)
    # print "removeList = %d" % len(removeList)

    return csa_new_files, new_files


def calc_senkei(kif_file):
    """ main routine of senkei calculation.
    """
    date = re.compile(
        ur'開始日時：([0-9]+)/([0-9]+)/([0-9]+) ([0-9]+):([0-9]+):([0-9]+)\r\n')
    sente = re.compile(ur'先手：(.+)\r\n')
    gote = re.compile(ur'後手：(.+)\r\n')
    senkei_line = re.compile(ur'戦型：(.+)\r\n')
    made = re.compile(ur'まで([0-9]+)手で([先後]手)の(入玉勝ち|勝ち)\r\n')
    summary = re.compile(
        ur"'summary:([a-z_ ]+):([^:]+)\ (lose|win|draw):([^:]+)\ (win|lose|draw)")

    reason_map = {u'illegal move': u'反則負け',
                  u'max_moves': u'最大手数',
                  u'oute_sennichite': u'王手千日手',
                  u'sennichite': u'千日手',
                  u'time up': u'時間切れ負け',
                  u'oute_kaihimore': u'王手回避もれ',
                  u'uchifuzume': u'打ち歩詰め'}
    num = 1
    senkei_table = {u'合計': 0}

    print(u','.join([u'年', u'月', u'日', u'時', u'分', u'秒', u'先手', u'後手', u'戦型',
                     u'手数', u'勝者(先後)', u'勝者', u'敗者', u'結果', u'棋譜ファイル', u'リンク'])
          + '\r').encode('cp932')

    records = []

    for k in kif_file:
        num += 1
        senkei_str = u'戦型データなし'
        root, ext = os.path.splitext(k)
        ext = ext
        csafile = root + '.csa'
        print 'analyze:' + str(num) + ':' + csafile
        k_file_obj = codecs.open(k, 'rU', 'cp932')
        try:
            lines = k_file_obj.readlines()
        except StandardError:
            print csafile + ' is ignored!!!!!!!!!!!!!!!!!!!!!!!!!!!'
            num -= 1
            k_file_obj.close()
            continue

        k_file_obj.close()
        #print li
        kifu = []
        broken = False
        for kif_line in lines:
            if kif_line.startswith(u'開'):
                ma_obj = date.match(kif_line)
                if ma_obj:
                    #print u'日時='.encode('utf-8') + repr(m.group(1,2,3,4,5,6))
                    kifu = list(ma_obj.group(1, 2, 3, 4, 5, 6))
                else:
                    kifu = list(2014, 1, 1, 0, 0,
                                0)  # dummy data (broken file...)
                    broken = True

            ma_obj2 = senkei_line.match(kif_line)
            if ma_obj2:
                #print u'戦型='.encode('utf-8') + m.group(1).encode('utf-8')
                senkei_str = ma_obj2.group(1)[:]
                senkei_table[u'合計'] += 1
                senkei_item = senkei_table.get(senkei_str, u'えらー')
                if senkei_item == u'えらー':
                    senkei_table[senkei_str] = 1
                else:
                    senkei_table[senkei_str] += 1

            ma_obj3 = sente.match(kif_line)
            if ma_obj3:
                #print u'先手='.encode('utf-8') + m.group(1).encode('utf-8')
                sente_str = ma_obj3.group(1)[:]

            ma_obj4 = gote.match(kif_line)
            if ma_obj4:
                #print u'後手='.encode('utf-8') + m.group(1).encode('utf-8')
                gote_str = ma_obj4.group(1)[:]

        if broken:
            print csafile + ' is ignored(broken file)!!!!!!!!!!!!!!!'
            num -= 1
            continue

        if senkei_str == u'戦型データなし':
            senkei_table[u'合計'] += 1
            item = senkei_table.get(senkei_str, u'えらー')
            if item == u'えらー':
                senkei_table[senkei_str] = 1
            else:
                senkei_table[senkei_str] += 1

        shouhai = False
        kifu += [sente_str]
        kifu += [gote_str]
        kifu += [senkei_str]

        if lines[len(lines) - 1].startswith(u'まで'):
            mat = made.match(lines[len(lines) - 1])
            if mat:
                kifu += [mat.group(1) + u'手']
                if mat.group(3) == u'入玉勝ち':
                    kachi = u'入玉勝ち'
                else:
                    kachi = u'投了'

                if mat.group(2) == u'先手':
                    kifu += [u'先手']
                    kifu += [sente_str]
                    kifu += [gote_str]
                    kifu += [kachi]
                else:
                    kifu += [u'後手']
                    kifu += [gote_str]
                    kifu += [sente_str]
                    kifu += [kachi]
                shouhai = True
            else:
                print u'xxxxx'.encode('cp932')
                print lines[len(lines) - 1].encode('cp932') + '\r'

        else:
            csafile2 = codecs.open(csafile, 'rU', 'cp932')
            li2 = csafile2.readlines()
            csafile2.close()

            kifu += [u'？手']

            for csa_line in li2:
                if csa_line.startswith("'summary:"):
                    m_summary = summary.match(csa_line)
                    if m_summary:
                        # print em
                        # print 'CSA '+m.group(1).encode('utf-8')
                        if m_summary.group(3) == u'draw':
                            kifu += [u'引き分け']
                            kifu += [u'']
                            kifu += [u'']
                            kifu += [reason_map[m_summary.group(1)]]
                            shouhai = True
                        elif m_summary.group(3) == u'win':
                            kifu += [u'先手']
                            kifu += [sente_str]
                            kifu += [gote_str]
                            kifu += [reason_map[m_summary.group(1)]]
                            shouhai = True
                        elif m_summary.group(5) == u'win':
                            kifu += [u'後手']
                            kifu += [gote_str]
                            kifu += [sente_str]
                            kifu += [reason_map[m_summary.group(1)]]
                            shouhai = True
                    else:
                        pass

            if shouhai is False:
                # print u'まで?手勝敗不明：不明'.encode('utf-8')
                kifu += [u'不明']
                kifu += [u'']
                kifu += [u'']
                kifu += [u'不明']

        kifu += [csafile[2:]]  # J
        kifu += [u'=HYPERLINK(J' + (u'%d' % num) + ')']  # K
        kifu += [u'=CELL("filename")']  # L
        unum = (u'%d' % num)
        kifu += [u'=CONCATENATE(LEFT(L' + unum + u',FIND("★",SUBSTITUTE(L' +
                 unum + u',"/","★",LEN(L' + unum + u')-LEN(SUBSTITUTE(L' + unum
                 + u',"/",""))),1)-1),"/",J' + unum + u')']  # M
        kifu += [u'=HYPERLINK(RIGHT(M' +
                 unum + u',LEN(M' + unum + u')-1))'] # N

        # kifu_t = u','.join(kifu)
        # print kifu_t.encode('utf-8')
        # print 'analyze:' + str(num) + ':' + csafile
        records += [kifu[:]]

    return records, senkei_table


def write_row(ws1, row_idx, reco):
    """ write kif record to excel sheet.
    """
    col = unicode(get_column_letter(1))
    row = unicode(str(row_idx))
    cell_name = u'%s%s' % (col, row)
    try:
        ws1[cell_name] = datetime.datetime(
            int(reco[0]), int(reco[1]), int(reco[2]), int(reco[3]),
            int(reco[4]), int(reco[5]))
    except StandardError:
        print reco[0:5]

    for i in range(6, len(reco)):
        col2 = unicode(get_column_letter(i + 1 - 5))
        cell_name2 = u'%s%s' % (col2, row)
        ws1[cell_name2] = reco[i]


def write_excel_xml(kif_record, dest_filename):
    """ write excel file.
    """
    wbook = Workbook()

    ws2 = wbook.active
    ws2.title = u'戦型一覧表'

    ws2.append([u'試合開始日時', u'先手', u'後手', u'戦型', u'手数', u'勝者(先後)', u'勝者', u'敗者',
                u'結果', u'棋譜ファイル', u'リンク(Excel)', u'ファイルパス1', u'ファイルパス2',
                u'リンク(LibreOffice)'])

    for row in range(2, len(kif_record) + 2):
        write_row(ws2, row, kif_record[row - 2])
        print row - 1

    wbook.save(filename=dest_filename)
    print "file=%s: %d records are saved(^^)." % (dest_filename,
                                                  len(kif_record))


if __name__ == '__main__':
    if len(sys.argv) != 2:
        print 'usage: python hikikaku.py sheetname'
    else:
        print sys.argv[1]
        CSA, KIF = retrieve_files(".")
        RECORD, SENKEI = calc_senkei(KIF)
        write_excel_xml(RECORD, sys.argv[1])
        for x in SENKEI.keys():
            print "%s: %d" % (x.encode('utf-8'), SENKEI[x])
